# Donovan Appelhans
# UWYO COSC 1010
# 10/14/2024
# HW 01
# Lab Section: 12
# Sources, people worked with, help given to:
# ChatGPT

import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mudkip Pixel Art"

cell_size = 5

for col in ws.columns:
    column = col[0].column_letter
    ws.column_dimensions[column].width = cell_size

for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = cell_size * 6

colors = {
    "blue": "4A90E2",
    "orange": "F5A623",
    "yellow": "F8E71C",
    "black": "000000",
    "white": "FFFFFF"
}

pixel_art = [
    ["blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue"],
    ["blue", "blue", "orange", "blue", "blue", "blue", "orange", "blue", "blue", "blue"],
    ["blue", "blue", "blue", "orange", "orange", "blue", "blue", "blue", "blue", "blue"],
    ["blue", "blue", "yellow", "blue", "blue", "yellow", "blue", "blue", "blue", "blue"],
    ["blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue"],
    ["blue", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "blue"],
    ["blue", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "yellow", "blue"],
    ["blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue", "blue"],
    ["blue", "blue", "black", "black", "black", "black", "black", "blue", "blue", "blue"],
    ["blue", "blue", "black", "black", "black", "black", "black", "blue", "blue", "blue"]
]

for row_idx, row in enumerate(pixel_art, start=1):
    for col_idx, color_name in enumerate(row, start=1):
        color = colors[color_name]
        cell = ws.cell(row=row_idx, column=col_idx)
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.fill = fill

wb.save("mudkip_pixel_art_10x10.xlsx")